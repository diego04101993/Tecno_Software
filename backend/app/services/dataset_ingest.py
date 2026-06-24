from __future__ import annotations

from datetime import date, datetime
from hashlib import md5
from io import BytesIO, StringIO
import csv
import re

from openpyxl import load_workbook

from app.models.entities import DatasetColumnType


SUPPORTED_DATASET_EXTENSIONS = {".csv", ".xlsx"}


def slugify_header(value: str, fallback_index: int) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", value.strip().lower())
    normalized = normalized.strip("_")
    return normalized or f"column_{fallback_index + 1}"


def make_unique_keys(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    keys: list[str] = []
    for index, header in enumerate(headers):
        base = slugify_header(header, index)
        counts[base] = counts.get(base, 0) + 1
        key = base if counts[base] == 1 else f"{base}_{counts[base]}"
        keys.append(key)
    return keys


def serialize_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return text or None


def infer_column_type(values: list[object]) -> DatasetColumnType:
    filtered = [value for value in values if value not in (None, "")]
    if not filtered:
        return DatasetColumnType.EMPTY
    if all(isinstance(value, bool) for value in filtered):
        return DatasetColumnType.BOOLEAN
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in filtered):
        return DatasetColumnType.NUMBER
    if all(isinstance(value, str) and looks_like_boolean(value) for value in filtered):
        return DatasetColumnType.BOOLEAN
    if all(isinstance(value, str) and looks_like_number(value) for value in filtered):
        return DatasetColumnType.NUMBER
    if all(isinstance(value, str) and looks_like_datetime(value) for value in filtered):
        return DatasetColumnType.DATETIME
    if all(isinstance(value, datetime | date) for value in filtered):
        return DatasetColumnType.DATETIME
    return DatasetColumnType.TEXT


def looks_like_number(value: str) -> bool:
    normalized = value.strip().replace(",", "")
    if not normalized:
        return False
    try:
        float(normalized)
        return True
    except ValueError:
        return False


def looks_like_boolean(value: str) -> bool:
    return value.strip().lower() in {"true", "false", "si", "no", "yes", "0", "1"}


def looks_like_datetime(value: str) -> bool:
    candidate = value.strip()
    if not candidate:
        return False
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%H:%M",
    ):
        try:
            datetime.strptime(candidate, fmt)
            return True
        except ValueError:
            continue
    return False


def decode_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


def extract_tabular_snapshot(file_name: str, mime_type: str | None, file_bytes: bytes) -> dict:
    lower_name = file_name.lower()
    if lower_name.endswith(".csv") or mime_type in {"text/csv", "application/csv"}:
        return parse_csv_snapshot(file_name, file_bytes)
    if lower_name.endswith(".xlsx") or mime_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }:
        return parse_xlsx_snapshot(file_name, file_bytes)
    raise ValueError("Formato no soportado. Solo se permiten archivos CSV o XLSX.")


def parse_csv_snapshot(file_name: str, file_bytes: bytes) -> dict:
    decoded = decode_csv_bytes(file_bytes)
    reader = csv.reader(StringIO(decoded))
    rows = list(reader)
    return build_snapshot(file_name=file_name, rows=rows, detected_sheet_name=None)


def parse_xlsx_snapshot(file_name: str, file_bytes: bytes) -> dict:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return build_snapshot(file_name=file_name, rows=rows, detected_sheet_name=sheet.title)


def build_snapshot(file_name: str, rows: list[list[object]], detected_sheet_name: str | None) -> dict:
    if not rows:
        raise ValueError("El archivo no contiene filas para importar.")

    header_row = rows[0]
    headers = [str(value).strip() if value not in (None, "") else f"Columna {index + 1}" for index, value in enumerate(header_row)]
    column_keys = make_unique_keys(headers)

    body_rows = rows[1:]
    normalized_rows: list[dict] = []
    raw_columns: dict[str, list[object]] = {key: [] for key in column_keys}

    next_row_index = 1
    for raw_row in body_rows:
        row_data: dict[str, object] = {}
        for index, key in enumerate(column_keys):
            serialized = serialize_cell(raw_row[index] if index < len(raw_row) else None)
            row_data[key] = serialized
            raw_columns[key].append(serialized)
        if all(value in (None, "") for value in row_data.values()):
            continue
        normalized_rows.append(
            {
                "row_index": next_row_index,
                "row_data_json": row_data,
                "row_hash": md5(str(sorted(row_data.items())).encode("utf-8")).hexdigest(),
            }
        )
        next_row_index += 1

    columns = []
    for index, key in enumerate(column_keys):
        samples = [value for value in raw_columns[key] if value not in (None, "")]
        sample = samples[0] if samples else None
        columns.append(
            {
                "column_key": key,
                "display_name": headers[index],
                "source_name": headers[index],
                "data_type": infer_column_type(raw_columns[key]),
                "position_index": index,
                "sample_value": str(sample) if sample is not None else None,
                "is_visible": True,
            }
        )

    return {
        "detected_sheet_name": detected_sheet_name,
        "column_count": len(columns),
        "row_count": len(normalized_rows),
        "columns": columns,
        "rows": normalized_rows,
        "summary_json": {
            "source_filename": file_name,
            "detected_sheet_name": detected_sheet_name,
            "column_names": headers,
            "preview_rows": min(20, len(normalized_rows)),
        },
    }
